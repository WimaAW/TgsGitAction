import os
import aiohttp
import asyncio
import openpyxl

async def request_weather(session, city, sheet, row):
    key = os.getenv('WEATHER_API_KEY')
    api = 'http://api.weatherapi.com/v1/current.json'
    payload = {
        "key": key,
        "q": city,
        "aqi": 'no'
    }
    async with session.get(api, params=payload, timeout=1000) as response:
        data = await response.json()
        print(data)
        write_to_excel(sheet, data, row)

def write_to_excel(sheet, data, row):
    sheet.cell(row=row, column=1, value=data.get('location', {}).get('name', ''))
    sheet.cell(row=row, column=2, value=data.get('current', {}).get('temp_c', ''))
    sheet.cell(row=row, column=3, value=data.get('current', {}).get('condition', {}).get('text', ''))
    sheet.cell(row=row, column=4, value=data.get('current', {}).get('wind_kph', ''))
    sheet.cell(row=row, column=5, value=data.get('current', {}).get('wind_dir', ''))
    sheet.cell(row=row, column=6, value=data.get('current', {}).get('humidity', ''))
    sheet.cell(row=row, column=7, value=data.get('current', {}).get('uv', ''))

async def main():
    file = 'Kecamatan.xlsx'
    output_file = 'output_file.xlsx'

    file_obj = openpyxl.load_workbook(file.strip())
    sheet_obj = file_obj.active

    max_row = sheet_obj.max_row

    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active

    headers = ['Kecamatan', 'Temperature (C)', 'Cuaca', 'Kecepatan Angin (kph)', 'Arah Angin', 'Kelembaban', 'Kadar UV']
    for col_num, header in enumerate(headers, 1):
        output_sheet.cell(row=1, column=col_num, value=header)

    async with aiohttp.ClientSession() as session:
        tasks = []
        for row in range(2, max_row + 1):
            kecamatan = sheet_obj.cell(row=row, column=2).value
            task = asyncio.create_task(request_weather(session, kecamatan, output_sheet, row))
            tasks.append(task)
        
        await asyncio.gather(*tasks)

    output_wb.save(output_file)

if __name__ == '__main__':
    asyncio.run(main())